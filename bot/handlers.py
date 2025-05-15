# Здесь будут хендлеры для пользовательских команд и сообщений 
import json
import os
from aiogram import types, F
from aiogram.filters import Command
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup, default_state
from .db import (
    create_user, update_user_profile, get_user_by_telegram_id, add_workout, confirm_payment, add_meal, update_last_active,
    get_user_workouts, get_user_meals, remove_payment_method_id
)
from .ai import ask_gpt, generate_workout_via_ai, analyze_food_photo_via_ai, generate_workout_via_ai_with_history
from aiogram.types import InlineKeyboardMarkup, InlineKeyboardButton, ReplyKeyboardMarkup, KeyboardButton, InputFile
from aiogram.types import BufferedInputFile
import io
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
import httpx
from .payments import create_payment_link
from datetime import datetime

SUBSCRIPTION_AMOUNT = os.getenv("SUBSCRIPTION_AMOUNT", "800")
MANAGER_NICK = os.getenv("MANAGER_NICK", "@your_manager")

MAIN_MENU = ReplyKeyboardMarkup(
    keyboard=[
        [KeyboardButton(text="Получить новую тренировку")],
        [KeyboardButton(text="Подсчет калорий")],
        [KeyboardButton(text="История")],
    ],
    resize_keyboard=True
)
MENU_BUTTONS = ["Получить новую тренировку", "Подсчет калорий", "История"]

ADMIN_MENU_BUTTON = KeyboardButton(text="Пуш-рассылка")

class ProfileStates(StatesGroup):
    goal = State()
    level = State()
    health_issues = State()
    location = State()
    workouts_per_week = State()
    height = State()
    weight = State()
    age = State()
    gender = State()

class CaloriesStates(StatesGroup):
    waiting_for_photo = State()

class PushStates(StatesGroup):
    waiting_for_text = State()
    waiting_for_audience = State()
    waiting_for_confirm = State()

WORKOUT_STATE = "workout_state"

async def mark_active(message):
    await update_last_active(message.from_user.id)

def is_positive_int(text):
    try:
        value = int(text)
        return value > 0
    except (ValueError, TypeError):
        return False

async def require_payment(message: types.Message, user=None):
    if user is None:
        user = await get_user_by_telegram_id(message.from_user.id)
    if not user or not user.get("is_paid"):
        await message.answer(f"Эта функция доступна только после оплаты подписки (стоимость {SUBSCRIPTION_AMOUNT}₽). Используй /pay для получения доступа.")
        return False
    # Новая проверка даты окончания подписки
    paid_until = user.get("paid_until")
    if paid_until:
        try:
            paid_until_date = datetime.fromisoformat(paid_until)
            if paid_until_date < datetime.utcnow():
                await message.answer("Ваша подписка истекла. Оформите новую для доступа к функциям с помощью /pay")
                return False
        except Exception:
            pass
    return True

# --- Хендлеры ---
from aiogram import Router
router = Router()

@router.message(Command("start"))
async def cmd_start(message: types.Message, state: FSMContext):
    await mark_active(message)
    try:
        user = await get_user_by_telegram_id(message.from_user.id)
        menu = await get_main_menu(message.from_user.id)
        if not user or not user.get("is_paid"):
            user = await create_user(
                telegram_id=message.from_user.id,
                username=message.from_user.username,
                first_name=message.from_user.first_name,
                last_name=message.from_user.last_name,
            )
            await message.answer(
                "Привет! Я фитнес-бот. Помогу тебе с тренировками, питанием и мотивацией!\n\n"
                "Давай начнем с небольшой анкеты.\n\n"
                "1. Какая у тебя цель? (Похудеть/Набрать массу/Поддерживать форму)"
            )
            await state.set_state(ProfileStates.goal)
            return
        await message.answer("Добро пожаловать! У тебя активна подписка. Используй меню ниже:", reply_markup=menu)
    except Exception as e:
        await message.answer("Произошла ошибка при обработке команды. Попробуйте позже.")
        print(f"Ошибка в /start: {e}")

@router.message(ProfileStates.goal)
async def process_goal(message: types.Message, state: FSMContext):
    try:
        await state.update_data(goal=message.text)
        await message.answer("2. Какой у тебя уровень подготовки? (Новичок/Средний/Продвинутый)")
        await state.set_state(ProfileStates.level)
    except Exception as e:
        await message.answer("Произошла ошибка. Попробуйте позже.")
        print(f"Ошибка в анкете (goal): {e}")

@router.message(ProfileStates.level)
async def process_level(message: types.Message, state: FSMContext):
    try:
        await state.update_data(level=message.text)
        await message.answer("3. Есть ли ограничения по здоровью?")
        await state.set_state(ProfileStates.health_issues)
    except Exception as e:
        await message.answer("Произошла ошибка. Попробуйте позже.")
        print(f"Ошибка в анкете (level): {e}")

@router.message(ProfileStates.health_issues)
async def process_health_issues(message: types.Message, state: FSMContext):
    try:
        await state.update_data(health_issues=message.text)
        await message.answer("4. Где планируете заниматься? (Дома/В зале/На улице)")
        await state.set_state(ProfileStates.location)
    except Exception as e:
        await message.answer("Произошла ошибка. Попробуйте позже.")
        print(f"Ошибка в анкете (health_issues): {e}")

@router.message(ProfileStates.location)
async def process_location(message: types.Message, state: FSMContext):
    try:
        await state.update_data(location=message.text)
        await message.answer("5. Сколько раз в неделю хотите тренироваться?")
        await state.set_state(ProfileStates.workouts_per_week)
    except Exception as e:
        await message.answer("Произошла ошибка. Попробуйте позже.")
        print(f"Ошибка в анкете (location): {e}")

@router.message(ProfileStates.workouts_per_week)
async def process_workouts_per_week(message: types.Message, state: FSMContext):
    try:
        await state.update_data(workouts_per_week=message.text)
        await message.answer("6. Рост (см)?")
        await state.set_state(ProfileStates.height)
    except Exception as e:
        await message.answer("Произошла ошибка. Попробуйте позже.")
        print(f"Ошибка в анкете (workouts_per_week): {e}")

@router.message(ProfileStates.height)
async def process_height(message: types.Message, state: FSMContext):
    try:
        await state.update_data(height=message.text)
        await message.answer("7. Вес (кг)?")
        await state.set_state(ProfileStates.weight)
    except Exception as e:
        await message.answer("Произошла ошибка. Попробуйте позже.")
        print(f"Ошибка в анкете (height): {e}")

@router.message(ProfileStates.weight)
async def process_weight(message: types.Message, state: FSMContext):
    try:
        await state.update_data(weight=message.text)
        await message.answer("8. Возраст?")
        await state.set_state(ProfileStates.age)
    except Exception as e:
        await message.answer("Произошла ошибка. Попробуйте позже.")
        print(f"Ошибка в анкете (weight): {e}")

@router.message(ProfileStates.age)
async def process_age(message: types.Message, state: FSMContext):
    try:
        await state.update_data(age=message.text)
        await message.answer("9. Пол (М/Ж)?")
        await state.set_state(ProfileStates.gender)
    except Exception as e:
        await message.answer("Произошла ошибка. Попробуйте позже.")
        print(f"Ошибка в анкете (age): {e}")

@router.message(ProfileStates.gender)
async def process_gender(message: types.Message, state: FSMContext):
    try:
        await state.update_data(gender=message.text)
        data = await state.get_data()
        # Формируем сводку анкеты
        summary = (
            f"Проверь, всё ли верно:\n"
            f"1. Цель: {data.get('goal', '')}\n"
            f"2. Уровень: {data.get('level', '')}\n"
            f"3. Ограничения: {data.get('health_issues', '')}\n"
            f"4. Где планируешь заниматься: {data.get('location', '')}\n"
            f"5. Сколько раз в неделю: {data.get('workouts_per_week', '')}\n"
            f"6. Рост: {data.get('height', '')}\n"
            f"7. Вес: {data.get('weight', '')}\n"
            f"8. Возраст: {data.get('age', '')}\n"
            f"9. Пол: {data.get('gender', '')}"
        )
        kb = InlineKeyboardMarkup(
            inline_keyboard=[
                [InlineKeyboardButton(text="✅ Всё верно", callback_data="profile_confirm")],
                [InlineKeyboardButton(text="Изменить ответы", callback_data="profile_restart")],
            ]
        )
        await message.answer(summary, reply_markup=kb)
        await state.set_state("profile_confirm_wait")
    except Exception as e:
        await message.answer("Произошла ошибка. Попробуйте позже.")
        print(f"Ошибка в анкете (gender): {e}")

@router.callback_query(lambda c: c.data in ["profile_confirm", "profile_restart"])
async def profile_confirm_callback(callback_query: types.CallbackQuery, state: FSMContext):
    try:
        current_state = await state.get_state()
        # Если состояние уже очищено, игнорируем повторный callback
        if not current_state or current_state != "profile_confirm_wait":
            await callback_query.answer("Анкета уже подтверждена или сброшена.", show_alert=False)
            return
        data = await state.get_data()
        if callback_query.data == "profile_confirm":
            await update_user_profile(
                telegram_id=callback_query.from_user.id,
                goal=data.get("goal"),
                level=data.get("level"),
                health_issues=data.get("health_issues"),
                location=data.get("location"),
                workouts_per_week=data.get("workouts_per_week"),
                height=data.get("height"),
                weight=data.get("weight"),
                age=data.get("age"),
                gender=data.get("gender"),
            )
            user = await get_user_by_telegram_id(callback_query.from_user.id)
            # Удаляем сообщение с кнопками, чтобы нельзя было кликнуть повторно
            try:
                await callback_query.message.delete()
            except Exception as e:
                print(f"Ошибка при удалении сообщения: {e}")
            await state.clear()
            # Сообщение о генерации плана
            if user.get("is_paid"):
                wait_msg = await callback_query.message.answer("Профиль подтверждён! Генерирую персональный план, пожалуйста, подождите...")
            else:
                wait_msg = await callback_query.message.answer("Профиль подтверждён! Генерирую бесплатную тренировку, пожалуйста, подождите...")
            workout_text = await generate_workout_via_ai(user)
            if not workout_text or not workout_text.strip():
                await callback_query.message.answer(
                    "Не удалось сгенерировать тренировку: описание отсутствует. Попробуйте ещё раз.",
                    reply_markup=MAIN_MENU
                )
                return
            # Проверяем статус оплаты
            if user.get("is_paid"):
                await add_workout(
                    user_id=user["id"],
                    workout_type="personal",
                    details=workout_text
                )
                await callback_query.message.answer(
                    "Профиль обновлён! Вот твой новый персональный план:\n\n" +
                    workout_text
                )
            else:
                await add_workout(
                    user_id=user["id"],
                    workout_type="free_trial",
                    details=workout_text
                )
                await callback_query.message.answer(
                    "Спасибо! Вот твоя первая бесплатная тренировка:\n\n" +
                    workout_text +
                    f"\n\nЕсли хочешь получить доступ к персональным тренировкам и другим функциям — оформи подписку! Стоимость подписки: {SUBSCRIPTION_AMOUNT}₽. /pay"
                )
            menu = await get_main_menu(callback_query.from_user.id)
            await callback_query.message.answer("Меню доступно ниже 👇", reply_markup=menu)
        else:
            try:
                await callback_query.message.delete()
            except Exception as e:
                print(f"Ошибка при удалении сообщения: {e}")
            await state.clear()
            await callback_query.message.answer(
                "Давай начнем заново!\n\n1. Какая у тебя цель? (Похудеть/Набрать массу/Поддерживать форму)"
            )
            await state.set_state(ProfileStates.goal)
        await callback_query.answer()
    except Exception as e:
        await callback_query.answer("Произошла ошибка при обработке анкеты.", show_alert=True)
        print(f"Ошибка в profile_confirm_callback: {e}")

@router.message(Command("pay"))
async def cmd_pay(message: types.Message):
    try:
        pay_url, payment_id = await create_payment_link(
            return_url=f"https://t.me/{os.getenv('BOT_USERNAME', 'your_bot_username')}",
            metadata={"telegram_id": str(message.from_user.id)}
        )
        await message.answer(
            f"Для оплаты подписки (стоимость {SUBSCRIPTION_AMOUNT}₽) перейдите по ссылке: {pay_url}\n\nПосле оплаты напишите администратору или дождитесь подтверждения."
        )
    except Exception as e:
        await message.answer("Ошибка при создании платежа. Попробуйте позже.")
        print(f"Ошибка в /pay: {e}")

@router.message(Command("confirm_payment"))
async def cmd_confirm_payment(message: types.Message):
    user = await get_user_by_telegram_id(message.from_user.id)
    if not user or user.get("role") != "admin":
        await message.answer("Нет доступа. Только для админов.")
        return
    parts = message.text.strip().split()
    if len(parts) != 2 or not parts[1].isdigit():
        await message.answer("Используй: /confirm_payment <telegram_id>")
        return
    telegram_id = int(parts[1])
    updated = await confirm_payment(telegram_id)
    if updated:
        await message.answer(f"Оплата подтверждена для пользователя {telegram_id}.")
    else:
        await message.answer("Пользователь не найден.")

@router.callback_query(lambda c: c.data == "pay_link")
async def process_pay_link(callback_query: types.CallbackQuery):
    pay_url = "https://yookassa.ru/pay/demo-link"
    await callback_query.message.answer(f"Для доступа ко всем функциям бота оплати подписку (стоимость {SUBSCRIPTION_AMOUNT}₽) по ссылке: {pay_url}\n\nПосле оплаты напиши администратору или дождись подтверждения.")
    await callback_query.answer()

@router.message(F.text == "Получить новую тренировку")
async def get_new_workout(message: types.Message, state: FSMContext):
    # Проверка на занятость
    data = await state.get_data()
    if data.get("is_busy"):
        await message.answer("Генерация уже выполняется, подождите...")
        return
    await state.update_data(is_busy=True)
    await state.clear()
    await mark_active(message)
    # Сразу даём отклик пользователю
    await message.answer("Команда принята, обрабатываю...")
    try:
        user = await get_user_by_telegram_id(message.from_user.id)
        if not await require_payment(message, user):
            await state.update_data(is_busy=False)
            return
        workout_text = await generate_workout_via_ai(user)
        if not workout_text or not workout_text.strip():
            await message.answer(
                "Не удалось сгенерировать тренировку: описание отсутствует. Попробуйте ещё раз.",
                reply_markup=MAIN_MENU
            )
            await state.update_data(is_busy=False)
            return
        # Сохраняем тренировку и историю в FSMContext
        await state.update_data(workout_text=workout_text, workout_history=[{"role": "user", "content": "Запрос тренировки"}, {"role": "assistant", "content": workout_text}], is_busy=False)
        # Кнопки
        kb = InlineKeyboardMarkup(
            inline_keyboard=[
                [InlineKeyboardButton(text="Выполнил", callback_data="workout_done")],
                [InlineKeyboardButton(text="Изменить тренировку", callback_data="workout_change")],
            ]
        )
        # Сохраняем тренировку с типом personal
        await add_workout(
            user_id=user["id"],
            workout_type="personal",
            details=workout_text
        )
        await message.answer(workout_text, reply_markup=kb)
    except Exception as e:
        await message.answer("Произошла ошибка при получении тренировки. Попробуйте позже.")
        print(f"Ошибка в get_new_workout: {e}")
        await state.update_data(is_busy=False)

@router.callback_query(lambda c: c.data == "workout_done")
async def workout_done_callback(callback_query: types.CallbackQuery, state: FSMContext):
    data = await state.get_data()
    workout_text = data.get("workout_text")
    if not workout_text or not workout_text.strip():
        await callback_query.message.answer(
            "Не удалось сохранить тренировку: описание отсутствует. Пожалуйста, запроси новую тренировку и попробуй снова.",
            reply_markup=MAIN_MENU
        )
        await state.clear()
        await callback_query.answer()
        return
    user = await get_user_by_telegram_id(callback_query.from_user.id)
    # Сохраняем только последнюю подтверждённую тренировку
    await add_workout(
        user_id=user["id"],
        workout_type="personal",
        details=workout_text
    )
    # Удаляем сообщение с кнопками
    try:
        await callback_query.message.delete()
    except Exception as e:
        print(f"Ошибка при удалении сообщения: {e}")
    await callback_query.message.answer("Молодец, так держать!", reply_markup=MAIN_MENU)
    await state.clear()
    await callback_query.answer()

@router.callback_query(lambda c: c.data == "workout_change")
async def workout_change_callback(callback_query: types.CallbackQuery, state: FSMContext):
    await callback_query.answer()
    data = await state.get_data()
    if data.get("is_busy"):
        await callback_query.message.answer("Генерация уже выполняется, подождите...")
        return
    await state.update_data(is_busy=True)
    # Системное сообщение сразу
    wait_msg = await callback_query.message.answer("Запрос принят, генерирую новую тренировку. Пожалуйста, подождите...")
    try:
        user = await get_user_by_telegram_id(callback_query.from_user.id)
        workout_history = data.get("workout_history", [])
        # Добавляем сообщение пользователя о недовольстве
        workout_history.append({"role": "user", "content": "Не нравится, давай другую тренировку."})
        # Формируем промпт для новой тренировки с историей
        new_workout_text = await generate_workout_via_ai_with_history(user, workout_history)
        # Обновляем историю
        workout_history.append({"role": "assistant", "content": new_workout_text})
        await state.update_data(workout_text=new_workout_text, workout_history=workout_history, is_busy=False)
        kb = InlineKeyboardMarkup(
            inline_keyboard=[
                [InlineKeyboardButton(text="Выполнил", callback_data="workout_done")],
                [InlineKeyboardButton(text="Изменить тренировку", callback_data="workout_change")],
            ]
        )
        # Удаляем старое сообщение с кнопками
        try:
            await callback_query.message.delete()
        except Exception as e:
            print(f"Ошибка при удалении сообщения: {e}")
        # Удаляем уведомление о генерации
        try:
            await wait_msg.delete()
        except Exception as e:
            print(f"Ошибка при удалении сообщения: {e}")
        # Отправляем новую тренировку с кнопками
        await callback_query.message.answer(new_workout_text, reply_markup=kb)
    except Exception as e:
        await callback_query.message.answer("Произошла ошибка при генерации новой тренировки. Попробуйте позже.")
        print(f"Ошибка в workout_change_callback: {e}")
        await state.update_data(is_busy=False)

@router.message(F.text == "Подсчет калорий")
async def start_calories(message: types.Message, state: FSMContext):
    # Проверка на занятость
    data = await state.get_data()
    if data.get("is_busy"):
        await message.answer("Анализ уже выполняется, подождите...")
        return
    await state.update_data(is_busy=True)
    await state.clear()
    await mark_active(message)
    try:
        user = await get_user_by_telegram_id(message.from_user.id)
        if not await require_payment(message, user):
            await state.update_data(is_busy=False)
            return
        await message.answer("Команда принята, обрабатываю...")
        await message.answer("Пожалуйста, отправь фото еды.")
        await state.set_state(CaloriesStates.waiting_for_photo)
        await state.update_data(is_busy=False)
    except Exception as e:
        await message.answer("Произошла ошибка при запуске сценария подсчета калорий. Попробуйте позже.")
        print(f"Ошибка в start_calories: {e}")
        await state.update_data(is_busy=False)

@router.message(CaloriesStates.waiting_for_photo, F.photo)
async def process_calories_photo(message: types.Message, state: FSMContext):
    # Проверка на занятость
    data = await state.get_data()
    if data.get("is_busy"):
        await message.answer("Анализ уже выполняется, подождите...")
        return
    await state.update_data(is_busy=True)
    await mark_active(message)
    try:
        await message.answer("Команда принята, анализирую фото...")
        user = await get_user_by_telegram_id(message.from_user.id)
        photo = message.photo[-1]
        file = await message.bot.get_file(photo.file_id)
        file_path = file.file_path
        TELEGRAM_TOKEN = os.getenv("TELEGRAM_TOKEN")
        file_url = f'https://api.telegram.org/file/bot{TELEGRAM_TOKEN}/{file_path}'
        gpt_response = await analyze_food_photo_via_ai(file_url)
        # Пытаемся найти JSON в ответе
        json_start = gpt_response.find('{')
        json_end = gpt_response.rfind('}')
        if json_start != -1 and json_end != -1 and json_end > json_start:
            try:
                data = json.loads(gpt_response[json_start:json_end+1])
                desc = data.get("description", "Фото еды")
                if not desc or not desc.strip():
                    await message.answer(
                        "Не удалось сохранить приём пищи: описание отсутствует. Попробуйте ещё раз.",
                        reply_markup=MAIN_MENU
                    )
                    await state.clear()
                    await state.update_data(is_busy=False)
                    return
                await add_meal(
                    user_id=user["id"],
                    description=desc,
                    calories=data.get("calories"),
                    proteins=data.get("proteins"),
                    fats=data.get("fats"),
                    carbs=data.get("carbs")
                )
                await message.answer(
                    f"Описание: {desc}\n"
                    f"Калории: {data.get('calories', '')}\n"
                    f"Б: {data.get('proteins', '—')} г, Ж: {data.get('fats', '—')} г, У: {data.get('carbs', '—')} г",
                    reply_markup=MAIN_MENU
                )
            except Exception as e:
                await message.answer("Ошибка при разборе ответа ИИ. Попробуйте позже.")
                print(f"Ошибка парсинга JSON из Vision: {e}")
        else:
            await message.answer(gpt_response, reply_markup=MAIN_MENU)
        await state.clear()
    except Exception as e:
        await message.answer("Произошла ошибка при обработке фото. Попробуйте позже.")
        print(f"Ошибка в process_calories_photo: {e}")
    finally:
        await state.update_data(is_busy=False)

@router.message(CaloriesStates.waiting_for_photo)
async def process_calories_not_photo(message: types.Message, state: FSMContext):
    print(f"DEBUG: process_calories_not_photo called, message.text = '{message.text}'")
    if message.text in MENU_BUTTONS:
        await state.clear()
        if message.text == "История":
            await show_history(message, state)
        elif message.text == "Получить новую тренировку":
            await get_new_workout(message, state)
        elif message.text == "Подсчет калорий":
            await start_calories(message, state)
        return
    await message.answer("Пожалуйста, отправь именно фото еды.")

@router.message(F.text == "История")
async def show_history(message: types.Message, state: FSMContext):
    print("DEBUG: show_history called, state cleared")
    # Проверка на занятость
    data = await state.get_data()
    if data.get("is_busy"):
        await message.answer("История уже формируется, подождите...")
        return
    await state.update_data(is_busy=True)
    # Системное сообщение сразу
    await message.answer("Формирую историю, пожалуйста, подождите...")
    await state.clear()
    await mark_active(message)
    menu = await get_main_menu(message.from_user.id)
    try:
        user = await get_user_by_telegram_id(message.from_user.id)
        if not await require_payment(message, user):
            await state.update_data(is_busy=False)
            return
        print("DEBUG: show_history - user and payment ok")
        workouts = await get_user_workouts(user["id"])
        meals = await get_user_meals(user["id"])
        print(f"DEBUG: show_history - workouts: {len(workouts)}, meals: {len(meals)}")
        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = "Тренировки"
        ws1.append(["Дата", "Тип", "Описание", "Калории (если есть)"])
        # Стили для заголовков
        header_font = Font(bold=True)
        align_left = Alignment(horizontal="left", vertical="top", wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        for cell in ws1[1]:
            cell.font = header_font
            cell.alignment = align_left
            cell.border = thin_border
        for w in workouts:
            ws1.append([
                w.get("date", ""),
                w.get("workout_type", ""),
                w.get("details", ""),
                w.get("calories_burned", "")
            ])
        # Стили для данных
        for row in ws1.iter_rows(min_row=2, max_row=ws1.max_row):
            for cell in row:
                cell.alignment = align_left
                cell.border = thin_border
        # Автоширина
        for col in ws1.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    cell_len = len(str(cell.value)) if cell.value else 0
                    if cell_len > max_length:
                        max_length = cell_len
                except:
                    pass
            ws1.column_dimensions[col_letter].width = max(12, min(max_length + 2, 50))
        ws2 = wb.create_sheet(title="Питание")
        ws2.append(["Дата", "Описание", "Калории", "Белки (г)", "Жиры (г)", "Углеводы (г)"])
        for cell in ws2[1]:
            cell.font = header_font
            cell.alignment = align_left
            cell.border = thin_border
        for m in meals:
            ws2.append([
                m.get("date", ""),
                m.get("description", ""),
                m.get("calories", ""),
                m.get("proteins", ""),
                m.get("fats", ""),
                m.get("carbs", "")
            ])
        for row in ws2.iter_rows(min_row=2, max_row=ws2.max_row):
            for cell in row:
                cell.alignment = align_left
                cell.border = thin_border
        for col in ws2.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    cell_len = len(str(cell.value)) if cell.value else 0
                    if cell_len > max_length:
                        max_length = cell_len
                except:
                    pass
            ws2.column_dimensions[col_letter].width = max(12, min(max_length + 2, 50))
        file_stream = io.BytesIO()
        wb.save(file_stream)
        file_stream.seek(0)
        excel_file = BufferedInputFile(file_stream.read(), filename="history.xlsx")
        print("DEBUG: show_history - Excel file ready, sending to user")
        await message.answer_document(
            excel_file,
            caption="Ваша история в формате Excel",
            reply_markup=menu
        )
    except Exception as e:
        await message.answer("Произошла ошибка при формировании Excel-файла. Попробуйте позже.", reply_markup=menu)
        print(f"Ошибка в show_history: {e}")
    finally:
        await state.update_data(is_busy=False)

@router.message(F.text == "Пуш-рассылка")
async def push_start(message: types.Message, state: FSMContext):
    if not await is_admin(message.from_user.id):
        await message.answer("Нет доступа.")
        return
    await message.answer("Введите текст пуш-уведомления:")
    await state.set_state(PushStates.waiting_for_text)

@router.message(PushStates.waiting_for_text)
async def push_text(message: types.Message, state: FSMContext):
    await state.update_data(push_text=message.text)
    kb = InlineKeyboardMarkup(
        inline_keyboard=[
            [InlineKeyboardButton(text="Всем пользователям", callback_data="push_audience_all")],
            [InlineKeyboardButton(text="Только платным", callback_data="push_audience_paid")],
            [InlineKeyboardButton(text="Только бесплатным", callback_data="push_audience_free")],
        ]
    )
    await message.answer("Кому отправить уведомление?", reply_markup=kb)
    await state.set_state(PushStates.waiting_for_audience)

@router.callback_query(lambda c: c.data.startswith("push_audience_"))
async def push_audience(callback_query: types.CallbackQuery, state: FSMContext):
    audience = callback_query.data.replace("push_audience_", "")
    await state.update_data(push_audience=audience)
    kb = InlineKeyboardMarkup(
        inline_keyboard=[
            [InlineKeyboardButton(text="Подтвердить", callback_data="push_confirm")],
            [InlineKeyboardButton(text="Отмена", callback_data="push_cancel")],
        ]
    )
    text = (await state.get_data()).get("push_text", "")
    audience_str = {
        "all": "всем пользователям",
        "paid": "только платным",
        "free": "только бесплатным"
    }[audience]
    await callback_query.message.answer(f"Текст: {text}\nКому: {audience_str}\n\nПодтвердить рассылку?", reply_markup=kb)
    await state.set_state(PushStates.waiting_for_confirm)
    await callback_query.answer()

@router.callback_query(lambda c: c.data == "push_cancel")
async def push_cancel(callback_query: types.CallbackQuery, state: FSMContext):
    await state.clear()
    await callback_query.message.answer("Рассылка отменена.")
    await callback_query.answer()

@router.callback_query(lambda c: c.data == "push_confirm")
async def push_confirm(callback_query: types.CallbackQuery, state: FSMContext):
    data = await state.get_data()
    if data.get("is_busy"):
        await callback_query.answer("Рассылка уже выполняется, подождите...", show_alert=True)
        return
    await state.update_data(is_busy=True)
    text = data.get("push_text", "")
    audience = data.get("push_audience", "all")
    await callback_query.message.answer("Рассылка начата...")
    await state.clear()
    # Получаем пользователей
    try:
        if audience == "all":
            users = await get_users_by_audience(None)
        else:
            users = await get_users_by_audience(audience)
        count = 0
        errors = 0
        from aiogram import Bot
        bot = callback_query.bot
        for user in users:
            try:
                await bot.send_message(user["telegram_id"], text)
                count += 1
            except Exception as e:
                errors += 1
                print(f"Ошибка отправки {user.get('telegram_id')}: {e}")
        await callback_query.message.answer(f"Рассылка завершена. Успешно: {count}, ошибок: {errors}")
    except Exception as e:
        await callback_query.message.answer("Произошла ошибка при рассылке.")
        print(f"Ошибка в push_confirm: {e}")
    finally:
        await state.update_data(is_busy=False)
    await callback_query.answer()

@router.message(PushStates.waiting_for_audience)
async def push_waiting_audience_message(message: types.Message, state: FSMContext):
    await message.answer("Пожалуйста, выберите вариант с помощью кнопок ниже.")

@router.message(PushStates.waiting_for_confirm)
async def push_waiting_confirm_message(message: types.Message, state: FSMContext):
    await message.answer("Пожалуйста, подтвердите или отмените рассылку с помощью кнопок ниже.")

async def is_admin(telegram_id):
    user = await get_user_by_telegram_id(telegram_id)
    return user and user.get("role") == "admin"

async def get_users_by_audience(audience):
    url = f"{os.getenv('SUPABASE_URL')}/rest/v1/users"
    headers = {
        "apikey": os.getenv('SUPABASE_KEY'),
        "Authorization": f"Bearer {os.getenv('SUPABASE_KEY')}",
    }
    params = {}
    if audience == "paid":
        params = {"is_paid": "eq.true"}
    elif audience == "free":
        params = {"is_paid": "eq.false"}
    async with httpx.AsyncClient() as client:
        resp = await client.get(url, headers=headers, params=params)
        return resp.json()

@router.message(Command("reset"))
async def cmd_reset(message: types.Message, state: FSMContext):
    try:
        # Сбросить только анкетные поля пользователя (goal, level, health_issues, location, workouts_per_week, height, weight, age, gender)
        await update_user_profile(
            telegram_id=message.from_user.id,
            goal=None,
            level=None,
            health_issues=None,
            location=None,
            workouts_per_week=None,
            height=None,
            weight=None,
            age=None,
            gender=None,
        )
        await state.clear()
        await message.answer(
            "Анкета сброшена. Давай начнем заново!\n\n1. Какая у тебя цель? (Похудеть/Набрать массу/Поддерживать форму)"
        )
        await state.set_state(ProfileStates.goal)
    except Exception as e:
        await message.answer(f"Произошла ошибка при сбросе анкеты: {e}")
        print(f"Ошибка в /reset: {repr(e)}")

@router.message(Command("manager"))
async def cmd_manager(message: types.Message):
    await message.answer(f"По всем вопросам обращайтесь к менеджеру: {MANAGER_NICK}")
    
@router.message(Command("help"), flags={"order": 0})
async def cmd_help(message: types.Message):
    print("DEBUG: cmd_help called", message.text, message.entities)
    text = (
        "ℹ️ <b>Доступные команды:</b>\n"
        "/manager — связаться с менеджером по вопросам оплаты, подписки и другим вопросам.\n"
        "/cancel_autopay — отключить автосписание (рекуррентную оплату). Подписка продолжит действовать до конца оплаченного периода, далее автосписания не будет.\n"
        "/reset — сбросить анкету и пройти опрос заново.\n"
        "/pay — оплатить подписку.\n"
        "/help — показать это сообщение.\n"
    )
    await message.answer(text, parse_mode="HTML") 

@router.message(Command("cancel_autopay"))
async def cmd_cancel_autopay(message: types.Message):
    user = await get_user_by_telegram_id(message.from_user.id)
    if not user or not user.get("payment_method_id"):
        await message.answer("У вас не подключено автосписание или вы не авторизованы.")
        return
    await remove_payment_method_id(message.from_user.id)
    await message.answer(
        "Автосписание успешно отключено. Ваша подписка будет действовать до конца оплаченного периода, далее продление возможно вручную через /pay."
    )

@router.message(F.text & ~F.text.in_(MENU_BUTTONS) & ~F.text.startswith("/"), default_state, flags={"order": 100})
async def universal_ai_handler(message: types.Message, state: FSMContext):
    await message.answer("Команда принята, думаю...")
    await mark_active(message)
    print("universal_ai_handler called")
    try:
        user = await get_user_by_telegram_id(message.from_user.id)
        if not await require_payment(message, user):
            return
        gpt_response = await ask_gpt("", message.text)
        # Пытаемся найти JSON в ответе
        json_start = gpt_response.find('{')
        json_end = gpt_response.rfind('}')
        saved = False
        if json_start != -1 and json_end != -1 and json_end > json_start:
            try:
                data = json.loads(gpt_response[json_start:json_end+1])
                if data.get("type") == "meal":
                    desc = data.get("description", "")
                    if not desc or not desc.strip():
                        await message.answer(
                            "Не удалось сохранить приём пищи: описание отсутствует. Попробуйте ещё раз.",
                            reply_markup=MAIN_MENU
                        )
                        return
                    await add_meal(
                        user_id=user["id"],
                        description=desc,
                        calories=data.get("calories"),
                        proteins=data.get("proteins"),
                        fats=data.get("fats"),
                        carbs=data.get("carbs")
                    )
                    await message.answer(
                        f"Записал приём пищи: {desc} ({data.get('calories', '')} ккал)\n"
                        f"Б: {data.get('proteins', '—')} г, Ж: {data.get('fats', '—')} г, У: {data.get('carbs', '—')} г"
                    )
                    saved = True
                elif data.get("type") == "workout":
                    workout_desc = data.get("description", "")
                    if not workout_desc or not workout_desc.strip():
                        await message.answer(
                            "Не удалось сохранить тренировку: описание отсутствует. Попробуйте ещё раз.",
                            reply_markup=MAIN_MENU
                        )
                        return
                    await add_workout(user_id=user["id"], workout_type=data.get("workout_type", "custom"), details=workout_desc, calories_burned=data.get("calories_burned"))
                    await message.answer("Записал тренировку: {}".format(workout_desc))
                    saved = True
            except Exception as e:
                print(f"Ошибка парсинга JSON из ответа GPT: {e}")
        # Отправляем сам ответ ИИ (без JSON)
        if saved:
            # Если был JSON, отправляем только текст до него (совет/комментарий)
            if json_start > 0:
                await message.answer(gpt_response[:json_start].strip())
        else:
            await message.answer(gpt_response)
    except Exception as e:
        await message.answer("Произошла ошибка при обработке запроса к ИИ. Попробуйте позже.")
        print(f"Ошибка в universal_ai_handler: {e}")

@router.message()
async def any_message_handler(message: types.Message, state: FSMContext):
    print(f"DEBUG: any_message_handler called, message.text = '{message.text}'")
    user = await get_user_by_telegram_id(message.from_user.id)
    menu = await get_main_menu(message.from_user.id)
    if user and not user.get("is_paid"):
        pay_keyboard = InlineKeyboardMarkup(
            inline_keyboard=[
                [InlineKeyboardButton(text=f"Оплатить подписку ({SUBSCRIPTION_AMOUNT}₽)", callback_data="pay_link")]
            ]
        )
        await message.answer(
            f"Чтобы получить доступ к персональным тренировкам и другим функциям, оплати подписку. Стоимость подписки: {SUBSCRIPTION_AMOUNT}₽. После оплаты напиши администратору или дождись подтверждения.",
            reply_markup=pay_keyboard
        )
        # Также показываем меню
        await message.answer("Меню доступно ниже 👇", reply_markup=menu)
        return
    if message.text in MENU_BUTTONS:
        return
    await message.answer("Используй команды для работы с ботом.", reply_markup=menu)


# Модифицируем MAIN_MENU для админа
async def get_main_menu(telegram_id):
    if await is_admin(telegram_id):
        return ReplyKeyboardMarkup(
            keyboard=[
                [KeyboardButton(text="Получить новую тренировку")],
                [KeyboardButton(text="Подсчет калорий")],
                [KeyboardButton(text="История")],
                [ADMIN_MENU_BUTTON],
            ],
            resize_keyboard=True
        )
    else:
        return MAIN_MENU